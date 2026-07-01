from pathlib import Path
from openpyxl import load_workbook
from app.parsers.base import BaseParser
from app.models import RawDocument, Page, Table


class XlsxParser(BaseParser):
    extensions = [".xlsx", ".xlsm"]

    def parse(self, file_path: Path) -> RawDocument:
        wb = load_workbook(file_path, data_only=True)
        pages = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if all(cell is None for cell in row):
                    continue
                if i == 0:
                    headers = [str(cell).strip() if cell is not None else "" for cell in row]
                else:
                    row_dict = {}
                    for j, cell in enumerate(row):
                        key = headers[j] if j < len(headers) else f"col_{j}"
                        row_dict[key] = str(cell).strip() if cell is not None else ""
                    rows.append(row_dict)

            if not headers and rows:
                headers = [f"col_{j}" for j in range(len(rows[0]))]

            text_lines = []
            if headers:
                text_lines.append(" | ".join(headers))
            for row in rows:
                text_lines.append(" | ".join(str(row.get(h, "")) for h in headers))

            table = Table(
                headers=headers,
                rows=rows,
                caption=sheet_name,
                page=len(pages) + 1,
            )

            pages.append(
                Page(
                    number=len(pages) + 1,
                    text=f"[Tabelle: {sheet_name}]\n" + "\n".join(text_lines),
                    tables=[table],
                )
            )

        wb.close()

        if not pages:
            pages.append(Page(number=1, text="(Leere Arbeitsmappe)"))

        return RawDocument(
            pages=pages,
            metadata={
                "source_file": file_path.name,
                "format": file_path.suffix.lower().lstrip("."),
                "page_count": len(pages),
                "sheet_count": len(wb.sheetnames),
            },
        )
